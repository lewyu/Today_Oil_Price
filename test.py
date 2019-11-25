#!/usr/bin/env python
# -*- coding: utf-8 -*-
# coding:utf-8
import codecs
import csv

import requests
from pyquery import PyQuery as pq
# 引入excel模块
import openpyxl

def get_page(url):
    """发起请求 获得源码"""
    r = requests.get(url)
    r.encoding = 'GBK'
    html = r.text
    return html


def parse(text):
    """解析数据 写入文件"""
    doc = pq(text)
    # 获得每一行的tr标签
    tds = doc('table  tr').items()

    for td in tds:  # 地区	92号汽油	95号汽油	98号汽油	0号柴油
        place = td.find('td:first-child').text()  # 地区
        gas92 = td.find('td:nth-child(2)').text()  # 92号汽油
        gas95 = td.find('td:nth-child(3)').text()  # 95号汽油
        gas98 = td.find('td:nth-child(4)').text()  # 98号汽油
        gas00 = td.find('td:nth-child(5)').text()  # 0号汽油

        with codecs.open('Today_gas_price.csv', 'a', 'utf_8_sig') as f:
            writer = csv.writer(f, dialect='excel')
            writer.writerow([place, gas92, gas95, gas98, gas00])

    print("写入完成")

#
# def write_excel_xlsx(path, sheet_name, value):
#     index = len(value)
#     workbook = openpyxl.Workbook()
#     sheet = workbook.active
#     sheet.title = sheet_name
#     for i in range(0, index):
#         for j in range(0, len(value[i])):
#             sheet.cell(row=i + 1, column=j + 1, value=str(value[i][j]))
#     workbook.save(path)
#     print("xlsx格式表格写入数据成功！")
#
#

# def read_excel_xlsx(path, sheet_name):
#     workbook = openpyxl.load_workbook(path)
#     # sheet = wb.get_sheet_by_name(sheet_name)这种方式已经弃用，不建议使用
#     sheet = workbook[sheet_name]
#     for row in sheet.rows:
#         for cell in row:
#             print(cell.value, "\t", end="")
#         print()


if __name__ == "__main__":
    # url = "http://www.zuihaodaxue.cn/zuihaodaxuepaiming2018.html"
    url = "http://oil.usd-cny.com/"
    text = get_page(url)
    print(text)
    parse(text)
